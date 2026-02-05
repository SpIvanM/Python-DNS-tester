import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from data.models import DomainConfig, DnsResolver, QueryResult, PerformanceStats, BlockingStats, CategorizedBlockingStats


class ExcelGenerator:
    """
    Generates the comprehensive Excel output report.
    """
    def __init__(self, output_filepath: str, all_domains: List[DomainConfig], all_resolvers: List[DnsResolver], query_store):
        self.output_filepath = output_filepath
        self.all_domains = sorted(all_domains, key=lambda d: d.name)  # Ensure consistent domain order
        self.all_resolvers = sorted(all_resolvers, key=lambda r: r.name)  # Ensure consistent resolver order
        self.query_store = query_store
        self.workbook = openpyxl.Workbook()
        self._setup_workbook()

    def _setup_workbook(self):
        # Remove default sheet created by openpyxl.Workbook()
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])

    def generate_report(self,
                        performance_stats_by_resolver: Dict[str, PerformanceStats],
                        blocking_stats_by_resolver: Dict[str, BlockingStats],
                        categorized_blocking_stats_by_resolver: Dict[str, List[CategorizedBlockingStats]],
                        blocked_useful_domains_by_resolver: Dict[str, List[str]],
                        blocked_useless_domains_by_resolver: Dict[str, List[str]],
                        passed_useless_domains_by_resolver: Dict[str, List[str]]):
        """
        Orchestrates the creation of the Excel workbook, including the matrix and detail sheets.
        """
        print("Generating Excel report...")
        self._create_matrix_sheet()

        for resolver in self.all_resolvers:
            self._create_resolver_detail_sheet(
                resolver=resolver,
                performance_stats=performance_stats_by_resolver.get(resolver.url, PerformanceStats(None, None, None, None)),
                blocking_stats=blocking_stats_by_resolver.get(resolver.url, BlockingStats(0, 0, 0, 0, 0.0)),
                categorized_blocking_stats=categorized_blocking_stats_by_resolver.get(resolver.url, []),
                blocked_useful_domains=blocked_useful_domains_by_resolver.get(resolver.url, []),
                blocked_useless_domains=blocked_useless_domains_by_resolver.get(resolver.url, []),
                passed_useless_domains=passed_useless_domains_by_resolver.get(resolver.url, [])
            )

        try:
            self.workbook.save(self.output_filepath)
            print(f"Excel report successfully saved to '{self.output_filepath}'")
        except Exception as e:
            print(f"Error saving Excel report: {e}")

    def _create_matrix_sheet(self):
        """
        Creates the "DNS Matrix" sheet in the Excel workbook.
        """
        ws = self.workbook.create_sheet(title="DNS Matrix")

        # Headers
        headers = ['DNS Resolver'] + [d.name for d in self.all_domains]
        ws.append(headers)

        # Apply header style
        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for col_idx, cell in enumerate(ws[1]):
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            cell.border = thin_border
            # Adjust column width based on content
            if col_idx == 0:
                ws.column_dimensions[get_column_letter(col_idx + 1)].width = 25
            else:
                ws.column_dimensions[get_column_letter(col_idx + 1)].width = max(len(headers[col_idx]), 10) + 2  # Min width 12

        # Data rows
        for r_idx, resolver in enumerate(self.all_resolvers):
            row_data = [resolver.name]
            for domain in self.all_domains:
                # Retrieve the specific query result for this resolver and domain
                results = self.query_store.get_results(resolver_url=resolver.url, domain_name=domain.name)

                # There should ideally be only one result per resolver+domain pair
                if results:
                    status = results[0].status
                    row_data.append('.' if status == 'Resolved' else 'X')
                else:
                    row_data.append('?')  # Should not happen if all queries were made

            ws.append(row_data)

            # Apply data row style
            for col_idx, cell in enumerate(ws[r_idx + 2]):  # +2 because row 1 is header, +1 for 0-indexing
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 0:  # Resolver name column
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    def _create_resolver_detail_sheet(self,
                                      resolver: DnsResolver,
                                      performance_stats: PerformanceStats,
                                      blocking_stats: BlockingStats,
                                      categorized_blocking_stats: List[CategorizedBlockingStats],
                                      blocked_useful_domains: List[str],
                                      blocked_useless_domains: List[str],
                                      passed_useless_domains: List[str]):
        """
        Creates a dedicated sheet for a single DNS resolver, detailing its statistics and lists.
        """
        # Sheet title cannot contain '/', ':', '?', '*', '[', ']', so sanitize resolver name
        sheet_title = resolver.name
        for char in '/\\?:*[]':
            sheet_title = sheet_title.replace(char, '_')
        if len(sheet_title) > 31:  # Excel sheet name limit
            sheet_title = sheet_title[:28] + '...'

        ws = self.workbook.create_sheet(title=sheet_title)

        # Styling
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Helper to write a section
        def write_section(title, data_rows, start_row):
            ws.cell(row=start_row, column=1, value=title).font = bold_font
            current_row = start_row + 1
            for row_data in data_rows:
                for col_idx, cell_value in enumerate(row_data):
                    cell = ws.cell(row=current_row, column=col_idx + 1, value=cell_value)
                    cell.border = thin_border
                    if col_idx == 0:  # Label column
                        cell.font = bold_font
                current_row += 1
            return current_row + 1  # Return next available row, with a blank line separator

        current_row = 1

        # Resolver Name
        ws.cell(row=current_row, column=1, value="DNS Resolver:").font = bold_font
        ws.cell(row=current_row, column=2, value=resolver.url)
        current_row += 2

        # Performance Statistics
        perf_data = [
            ["Min Latency (ms)", f"{performance_stats.min_latency_ms:.2f}" if performance_stats.min_latency_ms is not None else "N/A"],
            ["Max Latency (ms)", f"{performance_stats.max_latency_ms:.2f}" if performance_stats.max_latency_ms is not None else "N/A"],
            ["Median Latency (ms)", f"{performance_stats.median_latency_ms:.2f}" if performance_stats.median_latency_ms is not None else "N/A"],
            ["Avg Latency (ms)", f"{performance_stats.avg_latency_ms:.2f}" if performance_stats.avg_latency_ms is not None else "N/A"],
        ]
        current_row = write_section("Performance Statistics (Resolved Queries)", perf_data, current_row)

        # Error Rate Statistics
        error_rate = blocking_stats.error_queries / blocking_stats.total_queries * 100 if blocking_stats.total_queries > 0 else 0.0
        error_data = [
            ["Total Queries", blocking_stats.total_queries],
            ["Error Queries", blocking_stats.error_queries],
            ["Error Rate (%)", f"{error_rate:.2f}%"]
        ]
        current_row = write_section("Error Rate Statistics", error_data, current_row)

        # Overall Blocking Statistics
        overall_blocking_data = [
            ["Total Non-Error Queries", blocking_stats.resolved_queries + blocking_stats.blocked_queries],
            ["Resolved Queries", blocking_stats.resolved_queries],
            ["Blocked Queries", blocking_stats.blocked_queries],
            ["Overall Blocked (%)", f"{blocking_stats.overall_blocked_percentage:.2f}%"],
        ]
        current_row = write_section("Overall Blocking Statistics", overall_blocking_data, current_row)

        # Categorized Blocking Statistics
        cat_blocking_header = [["Category", "Total Non-Error", "Blocked Count", "Blocked (%)"]]
        cat_blocking_rows = []
        for stat in categorized_blocking_stats:
            cat_blocking_rows.append([
                stat.category,
                stat.total_non_error_in_category,
                stat.blocked_in_category,
                f"{stat.percentage_blocked_in_category:.2f}%"
            ])
        current_row = write_section("Categorized Blocking Statistics", cat_blocking_header + cat_blocking_rows, current_row)

        # List of Blocked Useful Domains
        current_row = write_section("Blocked Useful Domains", [[d] for d in blocked_useful_domains] if blocked_useful_domains else [["None"]], current_row)

        # List of Blocked Useless Domains
        current_row = write_section("Blocked Useless Domains", [[d] for d in blocked_useless_domains] if blocked_useless_domains else [["None"]], current_row)

        # List of Passed Useless Domains
        current_row = write_section("Passed Useless Domains", [[d] for d in passed_useless_domains] if passed_useless_domains else [["None"]], current_row)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:  # Cap max width
                adjusted_width = 50
            ws.column_dimensions[column].width = adjusted_width